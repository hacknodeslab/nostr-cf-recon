"""Extract relay URLs from a nostr.watch xlsx export.

Usage: extract_relays.py <xlsx> <out.txt> [--online] [--clearnet]
  --online    keep only in_rstate=True (active rotation)
  --clearnet  drop tor/i2p
"""
import sys
from openpyxl import load_workbook

args = sys.argv[1:]
flags = {a for a in args if a.startswith("--")}
pos = [a for a in args if not a.startswith("--")]
src, dst = pos[0], pos[1]
only_online = "--online" in flags
only_clearnet = "--clearnet" in flags

wb = load_workbook(src, read_only=True)
ws = wb.active
rows = ws.iter_rows(values_only=True)
header = list(next(rows))
url_idx = header.index("url")
rstate_idx = header.index("in_rstate")
net_idx = header.index("network")

seen = set()
with open(dst, "w") as f:
    for row in rows:
        url = row[url_idx]
        if not url:
            continue
        if only_online and not row[rstate_idx]:
            continue
        if only_clearnet and row[net_idx] != "clearnet":
            continue
        url = str(url).strip()
        if url in seen:
            continue
        seen.add(url)
        f.write(url + "\n")

print(f"wrote {len(seen)} urls -> {dst}  (online={only_online}, clearnet={only_clearnet})")
